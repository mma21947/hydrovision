from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from ordens.models import OrdemServico
from clientes.models import Cliente
from tecnicos.models import Tecnico
from django.db.models import Sum, Count, Avg, Max, Min, Q
from django.utils import timezone
from datetime import timedelta, datetime
import json
from django.core.serializers.json import DjangoJSONEncoder
# Novas importações para exportação (mais simples)
from django.http import HttpResponse
from django.template.loader import render_to_string, get_template
import csv
import tempfile
import os

# Para exportação de Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Para exportação de PDF
from xhtml2pdf import pisa
from django.conf import settings
from io import BytesIO

# Create your views here.

@login_required
def lista_relatorios(request):
    """Página inicial com links para todos os relatórios disponíveis"""
    return render(request, 'relatorios/lista_relatorios.html')

@login_required
def relatorio_por_cliente(request):
    """Relatório de ordens agrupadas por cliente"""
    clientes = Cliente.objects.annotate(
        total_ordens=Count('ordens'),
        valor_total=Sum('ordens__valor_total'),
        ticket_medio=Avg('ordens__valor_total'),
        valor_max=Max('ordens__valor_total'),
        ultima_ordem=Max('ordens__data_abertura')
    ).order_by('-total_ordens')
    
    # Calcular estatísticas gerais
    total_clientes = clientes.count()
    total_ordens = sum(cliente.total_ordens for cliente in clientes)
    total_valor = sum(float(cliente.valor_total or 0) for cliente in clientes)
    ticket_medio_geral = total_valor / total_ordens if total_ordens > 0 else 0
    
    # Verificar se é uma solicitação de exportação
    formato_exportacao = request.GET.get('export', None)
    if formato_exportacao:
        return exportar_relatorio(
            request, 
            'relatorios/relatorio_por_cliente.html', 
            {
                'clientes': clientes, 
                'titulo': 'Relatório por Cliente',
                'total_clientes': total_clientes,
                'total_ordens': total_ordens,
                'total_valor': total_valor,
                'ticket_medio_geral': ticket_medio_geral,
                'data_geracao': timezone.now()
            }, 
            f'relatorio_por_cliente.{formato_exportacao}',
            formato_exportacao,
            clientes=clientes
        )
    
    context = {
        'clientes': clientes,
        'titulo': 'Relatório por Cliente',
        'total_clientes': total_clientes,
        'total_ordens': total_ordens,
        'total_valor': total_valor,
        'ticket_medio_geral': ticket_medio_geral
    }
    return render(request, 'relatorios/relatorio_por_cliente.html', context)

@login_required
def relatorio_por_tecnico(request):
    """Relatório de ordens agrupadas por técnico"""
    tecnicos = Tecnico.objects.annotate(
        total_ordens=Count('ordens'),
        valor_total=Sum('ordens__valor_total'),
        ticket_medio=Avg('ordens__valor_total'),
        valor_max=Max('ordens__valor_total'),
        ultima_ordem=Max('ordens__data_abertura'),
        concluidas=Count('ordens', filter=Q(ordens__status='concluida'))
    ).order_by('-total_ordens')
    
    # Calcular estatísticas gerais
    total_tecnicos = tecnicos.count()
    total_ordens = sum(tecnico.total_ordens for tecnico in tecnicos)
    total_valor = sum(float(tecnico.valor_total or 0) for tecnico in tecnicos)
    ticket_medio_geral = total_valor / total_ordens if total_ordens > 0 else 0
    total_concluidas = sum(tecnico.concluidas for tecnico in tecnicos)
    taxa_conclusao_geral = (total_concluidas / total_ordens * 100) if total_ordens > 0 else 0
    
    # Verificar se é uma solicitação de exportação
    formato_exportacao = request.GET.get('export', None)
    if formato_exportacao:
        # Preparar dados para exportação - convertendo QuerySet em lista de dicionários
        tecnicos_export = []
        for tecnico in tecnicos:
            # Calcular ticket médio por técnico
            ticket_medio = float(tecnico.valor_total or 0) / tecnico.total_ordens if tecnico.total_ordens > 0 else 0
            # Calcular taxa de conclusão por técnico
            taxa_conclusao = (tecnico.concluidas / tecnico.total_ordens * 100) if tecnico.total_ordens > 0 else 0
            
            tecnicos_export.append({
                'nome': tecnico.nome_completo,
                'total_ordens': tecnico.total_ordens,
                'concluidas': tecnico.concluidas,
                'valor_total': float(tecnico.valor_total or 0),
                'ticket_medio': ticket_medio,
                'ultima_ordem': tecnico.ultima_ordem,
                'taxa_conclusao': taxa_conclusao
            })
        
        context_export = {
            'tecnicos': tecnicos_export,
            'titulo': 'Relatório por Técnico',
            'total_tecnicos': total_tecnicos,
            'total_ordens': total_ordens,
            'total_valor': total_valor,
            'ticket_medio_geral': ticket_medio_geral,
            'total_concluidas': total_concluidas,
            'taxa_conclusao_geral': taxa_conclusao_geral,
            'data_geracao': timezone.now()
        }
        
        # Selecionar o formato de exportação
        if formato_exportacao == 'pdf':
            return exportar_pdf(request, 'relatorios/relatorio_por_tecnico.html', context_export, f'relatorio_por_tecnico.pdf')
        elif formato_exportacao == 'xlsx':
            return exportar_excel_tecnico(request, tecnicos_export, total_tecnicos, total_ordens, total_valor, f'relatorio_por_tecnico.xlsx')
        elif formato_exportacao == 'csv':
            return exportar_csv_tecnico(request, tecnicos_export, total_tecnicos, total_ordens, total_valor, f'relatorio_por_tecnico.csv')
        else:
            # Formato não suportado
            return HttpResponse("Formato de exportação não suportado", status=400)
    
    context = {
        'tecnicos': tecnicos,
        'titulo': 'Relatório por Técnico',
        'total_tecnicos': total_tecnicos,
        'total_ordens': total_ordens,
        'total_valor': total_valor,
        'ticket_medio_geral': ticket_medio_geral,
        'total_concluidas': total_concluidas,
        'taxa_conclusao_geral': taxa_conclusao_geral
    }
    return render(request, 'relatorios/relatorio_por_tecnico.html', context)

@login_required
def relatorio_por_periodo(request):
    """Relatório de ordens por período (últimos 30 dias por padrão)"""
    data_inicial = request.GET.get('data_inicial', (timezone.now() - timedelta(days=30)).date())
    data_final = request.GET.get('data_final', timezone.now().date())
    
    if isinstance(data_inicial, str):
        data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
    if isinstance(data_final, str):
        data_final = datetime.strptime(data_final, '%Y-%m-%d').date()
    
    ordens = OrdemServico.objects.filter(
        data_abertura__range=[data_inicial, data_final]
    ).order_by('-data_abertura')
    
    # Calcular estatísticas do período
    total_ordens = ordens.count()
    valor_total = ordens.aggregate(total=Sum('valor_total'))['total'] or 0
    ticket_medio = valor_total / total_ordens if total_ordens > 0 else 0
    concluidas = ordens.filter(status='concluida').count()
    taxa_conclusao = (concluidas / total_ordens * 100) if total_ordens > 0 else 0
    
    # Clientes atendidos no período
    clientes_atendidos = Cliente.objects.filter(ordens__in=ordens).distinct().count()
    
    # Técnicos que trabalharam no período
    tecnicos_ativos = Tecnico.objects.filter(ordens__in=ordens).distinct().count()
    
    # Agrupamento por dia para gráfico
    dados_por_dia = OrdemServico.objects.filter(
        data_abertura__range=[data_inicial, data_final]
    ).values('data_abertura__date').annotate(
        total=Count('id'),
        valor=Sum('valor_total')
    ).order_by('data_abertura__date')
    
    # Converter data para string para serialização em JSON
    dados_para_grafico = []
    for item in dados_por_dia:
        dados_para_grafico.append({
            'data_abertura__date': item['data_abertura__date'].strftime('%Y-%m-%d'),
            'total': item['total'],
            'valor': float(item['valor'] or 0)
        })
    
    # Verificar se é uma solicitação de exportação
    formato_exportacao = request.GET.get('export', None)
    if formato_exportacao:
        context_export = {
            'ordens': ordens,
            'data_inicial': data_inicial,
            'data_final': data_final,
            'titulo': 'Relatório por Período',
            'total_ordens': total_ordens,
            'valor_total': valor_total,
            'ticket_medio': ticket_medio,
            'concluidas': concluidas,
            'taxa_conclusao_geral': taxa_conclusao,
            'clientes_atendidos': clientes_atendidos,
            'tecnicos_ativos': tecnicos_ativos,
            'data_geracao': timezone.now()
        }
        return exportar_relatorio(
            request, 
            'relatorios/relatorio_por_periodo.html', 
            context_export, 
            f'relatorio_por_periodo.{formato_exportacao}',
            formato_exportacao,
            ordens=ordens
        )
    
    context = {
        'ordens': ordens,
        'dados_por_dia': json.dumps(dados_para_grafico),
        'data_inicial': data_inicial,
        'data_final': data_final,
        'titulo': 'Relatório por Período',
        'total_ordens': total_ordens,
        'valor_total': valor_total,
        'ticket_medio': ticket_medio,
        'concluidas': concluidas,
        'taxa_conclusao': taxa_conclusao,
        'clientes_atendidos': clientes_atendidos,
        'tecnicos_ativos': tecnicos_ativos
    }
    return render(request, 'relatorios/relatorio_por_periodo.html', context)

@login_required
def relatorio_financeiro(request):
    """Relatório financeiro com resumo de faturamento"""
    # Faturamento mensal
    mes_atual = timezone.now().month
    ano_atual = timezone.now().year
    
    faturamento_mensal = OrdemServico.objects.filter(
        data_abertura__month=mes_atual,
        data_abertura__year=ano_atual,
        status='concluida'
    ).aggregate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    )
    
    # Faturamento anual
    faturamento_anual = OrdemServico.objects.filter(
        data_abertura__year=ano_atual,
        status='concluida'
    ).aggregate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    )
    
    # Calcular crescimento
    mes_anterior = mes_atual - 1 if mes_atual > 1 else 12
    ano_mes_anterior = ano_atual if mes_atual > 1 else ano_atual - 1
    
    faturamento_mes_anterior = OrdemServico.objects.filter(
        data_abertura__month=mes_anterior,
        data_abertura__year=ano_mes_anterior,
        status='concluida'
    ).aggregate(
        total=Sum('valor_total'),
        quantidade=Count('id')
    )
    
    # Calcular variação percentual
    valor_mes_anterior = float(faturamento_mes_anterior['total'] or 0)
    valor_mes_atual = float(faturamento_mensal['total'] or 0)
    
    if valor_mes_anterior > 0:
        variacao_percentual = ((valor_mes_atual - valor_mes_anterior) / valor_mes_anterior) * 100
    else:
        variacao_percentual = 100 if valor_mes_atual > 0 else 0
    
    # Dados para gráfico mensal
    dados_ultimos_6_meses = []
    for i in range(5, -1, -1):
        data_mes = timezone.now() - timedelta(days=30*i)
        mes = data_mes.month
        ano = data_mes.year
        dados_mes = OrdemServico.objects.filter(
            data_abertura__month=mes,
            data_abertura__year=ano,
            status='concluida'
        ).aggregate(
            total=Sum('valor_total'),
            quantidade=Count('id')
        )
        
        # Calcular ticket médio mensal
        ticket_medio = float(dados_mes['total'] or 0) / dados_mes['quantidade'] if dados_mes['quantidade'] else 0
        
        dados_ultimos_6_meses.append({
            'mes': data_mes.strftime('%B/%Y'),
            'total': float(dados_mes['total'] or 0),
            'quantidade': dados_mes['quantidade'] or 0,
            'ticket_medio': ticket_medio
        })
    
    # Top 5 clientes com maior faturamento
    top_clientes = Cliente.objects.annotate(
        total_faturado=Sum('ordens__valor_total', filter=Q(ordens__status='concluida'))
    ).order_by('-total_faturado')[:5]
    
    # Top 5 técnicos com maior faturamento
    top_tecnicos = Tecnico.objects.annotate(
        total_faturado=Sum('ordens__valor_total', filter=Q(ordens__status='concluida'))
    ).order_by('-total_faturado')[:5]
    
    # Ticket médio anual
    ticket_medio_anual = float(faturamento_anual['total'] or 0) / faturamento_anual['quantidade'] if faturamento_anual['quantidade'] else 0
    
    # Verificar se é uma solicitação de exportação
    formato_exportacao = request.GET.get('export', None)
    if formato_exportacao:
        # Dados específicos para o relatório financeiro em Excel
        dados_exportacao = {
            'faturamento_mensal': faturamento_mensal,
            'faturamento_anual': faturamento_anual,
            'dados_ultimos_6_meses': dados_ultimos_6_meses,
            'titulo': 'Relatório Financeiro',
            'variacao_percentual': variacao_percentual,
            'top_clientes': top_clientes,
            'top_tecnicos': top_tecnicos,
            'ticket_medio_anual': ticket_medio_anual,
            'data_geracao': timezone.now()
        }
        
        return exportar_relatorio(
            request, 
            'relatorios/relatorio_financeiro.html', 
            dados_exportacao, 
            f'relatorio_financeiro.{formato_exportacao}',
            formato_exportacao,
            dados_meses=dados_ultimos_6_meses,
            faturamento_mensal=faturamento_mensal,
            faturamento_anual=faturamento_anual,
            top_clientes=top_clientes,
            top_tecnicos=top_tecnicos
        )
    
    context = {
        'faturamento_mensal': faturamento_mensal,
        'faturamento_anual': faturamento_anual,
        'dados_ultimos_6_meses': json.dumps(dados_ultimos_6_meses),
        'titulo': 'Relatório Financeiro',
        'variacao_percentual': variacao_percentual,
        'top_clientes': top_clientes,
        'top_tecnicos': top_tecnicos,
        'ticket_medio_anual': ticket_medio_anual
    }
    return render(request, 'relatorios/relatorio_financeiro.html', context)

# Função para exportação de relatórios
def exportar_relatorio(request, template_name, context, filename, formato, **data_objects):
    """
    Função para exportar relatórios em diferentes formatos
    """
    if formato == 'pdf':
        return exportar_pdf(request, template_name, context, filename)
    elif formato == 'xlsx':
        return exportar_excel(request, data_objects, filename)
    elif formato == 'csv':
        return exportar_csv(request, data_objects, filename)
    else:
        # Formato não suportado
        return HttpResponse("Formato de exportação não suportado", status=400)

def exportar_pdf(request, template_name, context, filename):
    """
    Exporta o relatório em formato PDF usando xhtml2pdf
    """
    # Determinar o template PDF correto com base no tipo de relatório
    if 'relatorio_por_cliente' in template_name:
        template_path = 'relatorios/pdf/relatorio_por_cliente.html'
    elif 'relatorio_por_tecnico' in template_name:
        template_path = 'relatorios/pdf/relatorio_por_tecnico.html'
    elif 'relatorio_por_periodo' in template_name:
        template_path = 'relatorios/pdf/relatorio_por_periodo.html'
    elif 'relatorio_financeiro' in template_name:
        template_path = 'relatorios/pdf/relatorio_financeiro.html'
    else:
        template_path = template_name
    
    # Renderizar o template
    template = get_template(template_path)
    html = template.render(context)
    
    # Criar resposta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Gerar o PDF
    pdf_status = pisa.CreatePDF(
        html,
        dest=response,
        encoding='utf-8'
    )
    
    # Se houve erro
    if pdf_status.err:
        return HttpResponse('Erro ao gerar o PDF: ' + str(pdf_status.err), content_type='text/plain')
    
    return response

def exportar_excel(request, data_objects, filename):
    """
    Exporta o relatório em formato Excel usando openpyxl
    """
    # Criar um novo workbook e selecionar a planilha ativa
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Relatório'
    
    # Estilos para o cabeçalho
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borders = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Estilos para células de dados
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_currency_alignment = Alignment(horizontal='right', vertical='center')
    
    # Definir headers com valor padrão
    headers = []
    
    # Decide que tipo de relatório exportar
    if 'clientes' in data_objects:
        # Título do relatório
        worksheet['A1'] = 'Relatório por Cliente'
        worksheet['A1'].font = Font(name='Arial', bold=True, size=14)
        worksheet.merge_cells('A1:E1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Data de geração
        worksheet['A2'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A2'].font = Font(name='Arial', italic=True)
        worksheet.merge_cells('A2:E2')
        
        # Cabeçalhos da tabela
        headers = ['Cliente', 'Total de Ordens', 'Valor Total (R$)', 'Ticket Médio (R$)', 'Última Ordem']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = borders
        
        # Dados
        row_num = 5
        for cliente in data_objects['clientes']:
            # Cliente
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = cliente.nome
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Total de ordens
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = cliente.total_ordens
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Valor total
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = float(cliente.valor_total or 0)
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            # Ticket médio
            cell = worksheet.cell(row=row_num, column=4)
            ticket_medio = float(cliente.valor_total or 0) / cliente.total_ordens if cliente.total_ordens > 0 else 0
            cell.value = ticket_medio
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            # Última ordem
            cell = worksheet.cell(row=row_num, column=5)
            cell.value = cliente.ultima_ordem.strftime('%d/%m/%Y') if cliente.ultima_ordem else '-'
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            row_num += 1
        
        # Resumo
        row_num += 2
        worksheet.cell(row=row_num, column=1).value = 'Resumo'
        worksheet.cell(row=row_num, column=1).font = Font(name='Arial', bold=True)
        worksheet.merge_cells(f'A{row_num}:E{row_num}')
        
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = 'Total de Clientes'
        worksheet.cell(row=row_num, column=2).value = len(data_objects['clientes'])
        
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = 'Total de Ordens'
        worksheet.cell(row=row_num, column=2).value = sum(cliente.total_ordens for cliente in data_objects['clientes'])
        
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = 'Valor Total'
        worksheet.cell(row=row_num, column=2).value = sum(float(cliente.valor_total or 0) for cliente in data_objects['clientes'])
        worksheet.cell(row=row_num, column=2).number_format = 'R$ #,##0.00'
    elif 'tecnicos' in data_objects:
        # Título do relatório
        worksheet['A1'] = 'Relatório por Técnico'
        worksheet['A1'].font = Font(name='Arial', bold=True, size=14)
        worksheet.merge_cells('A1:E1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Data de geração
        worksheet['A2'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A2'].font = Font(name='Arial', italic=True)
        worksheet.merge_cells('A2:E2')
        
        # Cabeçalhos da tabela
        headers = ['Técnico', 'Total de Ordens', 'Ordens Concluídas', 'Valor Total (R$)', 'Ticket Médio (R$)', 'Última Ordem', 'Taxa de Conclusão (%)']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = borders
        
        # Dados
        row_num = 5
        for tecnico in data_objects['tecnicos']:
            # Verificar se estamos lidando com um dicionário ou objeto modelo
            if isinstance(tecnico, dict):
                nome = tecnico['nome']
                total_ordens = tecnico['total_ordens']
                concluidas = tecnico['concluidas']
                valor_total = tecnico['valor_total']
                ticket_medio = tecnico['ticket_medio']
                ultima_ordem = tecnico['ultima_ordem']
            else:
                nome = tecnico.nome_completo
                total_ordens = tecnico.total_ordens
                concluidas = tecnico.concluidas
                valor_total = float(tecnico.valor_total or 0)
                ticket_medio = float(tecnico.ticket_medio or 0)
                ultima_ordem = tecnico.ultima_ordem
            
            # Nome do técnico
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = nome
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Total de ordens
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = total_ordens
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Ordens concluídas
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = concluidas
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Valor total
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = valor_total
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            # Ticket médio
            cell = worksheet.cell(row=row_num, column=5)
            cell.value = ticket_medio
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            # Última ordem
            cell = worksheet.cell(row=row_num, column=6)
            if isinstance(ultima_ordem, datetime):
                cell.value = ultima_ordem.strftime('%d/%m/%Y')
            else:
                cell.value = ultima_ordem.strftime('%d/%m/%Y') if ultima_ordem else '-'
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Taxa de conclusão
            cell = worksheet.cell(row=row_num, column=7)
            taxa_conclusao = (concluidas / total_ordens * 100) if total_ordens > 0 else 0
            cell.value = f"{taxa_conclusao:.1f}%"
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            row_num += 1
    elif 'ordens' in data_objects:
        # Relatório por período
        worksheet['A1'] = 'Relatório por Período'
        worksheet['A1'].font = Font(name='Arial', bold=True, size=14)
        worksheet.merge_cells('A1:G1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Data de geração
        worksheet['A2'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A2'].font = Font(name='Arial', italic=True)
        worksheet.merge_cells('A2:G2')
        
        # Cabeçalhos da tabela
        headers = ['ID', 'Cliente', 'Técnico', 'Data Abertura', 'Data Conclusão', 'Status', 'Valor Total (R$)']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = borders
        
        # Dados
        row_num = 5
        for ordem in data_objects['ordens']:
            # ID da ordem
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = ordem.id
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Cliente
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = ordem.cliente.nome
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Técnico
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = ordem.tecnico.nome_completo if ordem.tecnico else "-"
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Data Abertura
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = ordem.data_abertura.strftime('%d/%m/%Y')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Data Conclusão
            cell = worksheet.cell(row=row_num, column=5)
            cell.value = ordem.data_conclusao.strftime('%d/%m/%Y') if ordem.data_conclusao else "-"
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Status
            cell = worksheet.cell(row=row_num, column=6)
            status_map = {
                'concluida': 'Concluída',
                'em_andamento': 'Em Andamento',
                'aguardando': 'Aguardando',
                'cancelada': 'Cancelada'
            }
            cell.value = status_map.get(ordem.status, ordem.status)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Valor Total
            cell = worksheet.cell(row=row_num, column=7)
            cell.value = float(ordem.valor_total or 0)
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            row_num += 1
    elif 'dados_meses' in data_objects:
        # Relatório financeiro
        worksheet['A1'] = 'Relatório Financeiro'
        worksheet['A1'].font = Font(name='Arial', bold=True, size=14)
        worksheet.merge_cells('A1:D1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Data de geração
        worksheet['A2'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A2'].font = Font(name='Arial', italic=True)
        worksheet.merge_cells('A2:D2')
        
        # Cabeçalhos faturamento mensal
        worksheet['A4'] = 'Faturamento dos Últimos 6 Meses'
        worksheet['A4'].font = Font(name='Arial', bold=True)
        worksheet.merge_cells('A4:D4')
        
        headers = ['Mês/Ano', 'Quantidade de Ordens', 'Faturamento Total (R$)', 'Ticket Médio (R$)']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = borders
        
        # Dados dos meses
        row_num = 6
        for mes_data in data_objects['dados_meses']:
            # Mês/Ano
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = mes_data['mes']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Quantidade
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = mes_data['quantidade']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Faturamento Total
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = mes_data['total']
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            # Ticket Médio
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = mes_data['ticket_medio']
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            row_num += 1
            
        # Adicionar seção de Top Clientes
        row_num += 2
        worksheet.cell(row=row_num, column=1).value = 'Top 5 Clientes'
        worksheet.cell(row=row_num, column=1).font = Font(name='Arial', bold=True)
        worksheet.merge_cells(f'A{row_num}:C{row_num}')
        
        row_num += 1
        # Cabeçalhos top clientes
        top_headers = ['Cliente', 'Total Faturado (R$)']
        for col_num, header in enumerate(top_headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = borders
            
        # Dados dos top clientes
        row_num += 1
        for cliente in data_objects['top_clientes']:
            # Cliente
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = cliente.nome
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = borders
            
            # Total Faturado
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = float(cliente.total_faturado or 0)
            cell.font = data_font
            cell.alignment = data_currency_alignment
            cell.border = borders
            cell.number_format = 'R$ #,##0.00'
            
            row_num += 1
    else:
        # Relatório genérico (fallback)
        worksheet['A1'] = 'Relatório'
        worksheet['A1'].font = Font(name='Arial', bold=True, size=14)
        worksheet.merge_cells('A1:C1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['Sem dados disponíveis']
        worksheet['A3'] = 'Sem dados disponíveis para exportação'
    
    # Ajustar larguras das colunas (só se headers não estiver vazio)
    if headers:
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20
    
    # Configurar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar o arquivo
    workbook.save(response)
    
    return response

def exportar_csv(request, data_objects, filename):
    """
    Exporta o relatório em formato CSV
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Decide que tipo de relatório exportar
    if 'clientes' in data_objects:
        # Relatório por cliente
        writer.writerow(['Cliente', 'Total de Ordens', 'Valor Total (R$)', 'Ticket Médio (R$)', 'Última Ordem'])
        
        for cliente in data_objects['clientes']:
            ticket_medio = float(cliente.valor_total or 0) / cliente.total_ordens if cliente.total_ordens > 0 else 0
            writer.writerow([
                cliente.nome,
                cliente.total_ordens,
                f"{float(cliente.valor_total or 0):.2f}".replace('.', ','),
                f"{ticket_medio:.2f}".replace('.', ','),
                cliente.ultima_ordem.strftime('%d/%m/%Y') if cliente.ultima_ordem else '-'
            ])
        
        # Adicionar resumo
        writer.writerow([])
        writer.writerow(['Resumo', '', '', '', ''])
        writer.writerow(['Total de Clientes', len(data_objects['clientes']), '', '', ''])
        writer.writerow(['Total de Ordens', sum(cliente.total_ordens for cliente in data_objects['clientes']), '', '', ''])
        writer.writerow(['Valor Total (R$)', f"{sum(float(cliente.valor_total or 0) for cliente in data_objects['clientes']):.2f}".replace('.', ','), '', '', ''])
    elif 'tecnicos' in data_objects:
        # Relatório por técnico
        writer.writerow(['Técnico', 'Total de Ordens', 'Ordens Concluídas', 'Valor Total (R$)', 'Ticket Médio (R$)', 'Última Ordem', 'Taxa de Conclusão (%)'])
        
        for tecnico in data_objects['tecnicos']:
            # Verificar se estamos lidando com um dicionário ou objeto modelo
            if isinstance(tecnico, dict):
                nome = tecnico['nome']
                total_ordens = tecnico['total_ordens']
                concluidas = tecnico['concluidas']
                valor_total = tecnico['valor_total']
                ticket_medio = tecnico['ticket_medio']
                ultima_ordem = tecnico['ultima_ordem']
            else:
                nome = tecnico.nome_completo
                total_ordens = tecnico.total_ordens
                concluidas = tecnico.concluidas
                valor_total = float(tecnico.valor_total or 0)
                ticket_medio = float(tecnico.ticket_medio or 0)
                ultima_ordem = tecnico.ultima_ordem
            
            # Nome do técnico
            writer.writerow([
                nome,
                total_ordens,
                concluidas,
                f"{valor_total:.2f}".replace('.', ','),
                f"{ticket_medio:.2f}".replace('.', ','),
                ultima_ordem.strftime('%d/%m/%Y') if ultima_ordem else '-',
                f"{((concluidas / total_ordens) * 100):.1f}%"
            ])
        
        # Adicionar resumo
        writer.writerow([])
        writer.writerow(['Resumo', '', '', '', '', '', ''])
        writer.writerow(['Total de Técnicos', len(data_objects['tecnicos']), '', '', '', '', ''])
        writer.writerow(['Total de Ordens', sum(tecnico['total_ordens'] for tecnico in data_objects['tecnicos']), '', '', '', '', ''])
        writer.writerow(['Valor Total (R$)', f"{sum(tecnico['valor_total'] for tecnico in data_objects['tecnicos']):.2f}".replace('.', ','), '', '', '', '', ''])
    
    return response

def exportar_excel_tecnico(request, tecnicos, total_tecnicos, total_ordens, total_valor, filename):
    """
    Exportação específica para o relatório de técnicos em Excel
    """
    # Criar um novo workbook e selecionar a planilha ativa
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Relatório de Técnicos'
    
    # Estilos para o cabeçalho
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borders = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Estilos para células de dados
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_currency_alignment = Alignment(horizontal='right', vertical='center')
    
    # Título do relatório
    worksheet['A1'] = 'Relatório por Técnico'
    worksheet['A1'].font = Font(name='Arial', bold=True, size=14)
    worksheet.merge_cells('A1:G1')
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    # Data de geração
    worksheet['A2'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    worksheet['A2'].font = Font(name='Arial', italic=True)
    worksheet.merge_cells('A2:G2')
    
    # Cabeçalhos da tabela
    headers = ['Técnico', 'Total de Ordens', 'Ordens Concluídas', 'Valor Total (R$)', 'Ticket Médio (R$)', 'Última Ordem', 'Taxa de Conclusão (%)']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = borders
    
    # Dados
    row_num = 5
    for tecnico in tecnicos:
        # Nome do técnico
        cell = worksheet.cell(row=row_num, column=1)
        cell.value = tecnico['nome']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = borders
        
        # Total de ordens
        cell = worksheet.cell(row=row_num, column=2)
        cell.value = tecnico['total_ordens']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = borders
        
        # Ordens concluídas
        cell = worksheet.cell(row=row_num, column=3)
        cell.value = tecnico['concluidas']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = borders
        
        # Valor total
        cell = worksheet.cell(row=row_num, column=4)
        cell.value = tecnico['valor_total']
        cell.font = data_font
        cell.alignment = data_currency_alignment
        cell.border = borders
        cell.number_format = 'R$ #,##0.00'
        
        # Ticket médio
        cell = worksheet.cell(row=row_num, column=5)
        cell.value = tecnico['ticket_medio']
        cell.font = data_font
        cell.alignment = data_currency_alignment
        cell.border = borders
        cell.number_format = 'R$ #,##0.00'
        
        # Última ordem
        cell = worksheet.cell(row=row_num, column=6)
        ultima_ordem = tecnico['ultima_ordem']
        if ultima_ordem:
            cell.value = ultima_ordem.strftime('%d/%m/%Y')
        else:
            cell.value = '-'
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = borders
        
        # Taxa de conclusão
        cell = worksheet.cell(row=row_num, column=7)
        total_ordens = tecnico['total_ordens']
        concluidas = tecnico['concluidas']
        taxa_conclusao = (concluidas / total_ordens * 100) if total_ordens > 0 else 0
        cell.value = f"{taxa_conclusao:.1f}%"
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = borders
        
        row_num += 1
    
    # Resumo
    row_num += 2
    worksheet.cell(row=row_num, column=1).value = 'Resumo'
    worksheet.cell(row=row_num, column=1).font = Font(name='Arial', bold=True)
    worksheet.merge_cells(f'A{row_num}:G{row_num}')
    
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = 'Total de Técnicos'
    worksheet.cell(row=row_num, column=2).value = total_tecnicos
    
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = 'Total de Ordens'
    worksheet.cell(row=row_num, column=2).value = total_ordens
    
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = 'Valor Total'
    worksheet.cell(row=row_num, column=2).value = total_valor
    worksheet.cell(row=row_num, column=2).number_format = 'R$ #,##0.00'
    
    # Ajustar larguras das colunas
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20
    
    # Configurar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar o arquivo
    workbook.save(response)
    
    return response

def exportar_csv_tecnico(request, tecnicos, total_tecnicos, total_ordens, total_valor, filename):
    """
    Exportação específica para o relatório de técnicos em CSV
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Cabeçalhos
    writer.writerow(['Técnico', 'Total de Ordens', 'Ordens Concluídas', 'Valor Total (R$)', 'Ticket Médio (R$)', 'Última Ordem', 'Taxa de Conclusão (%)'])
    
    # Dados
    for tecnico in tecnicos:
        total_ordens = tecnico['total_ordens']
        concluidas = tecnico['concluidas']
        taxa_conclusao = (concluidas / total_ordens * 100) if total_ordens > 0 else 0
        
        writer.writerow([
            tecnico['nome'],
            tecnico['total_ordens'],
            tecnico['concluidas'],
            f"{tecnico['valor_total']:.2f}".replace('.', ','),
            f"{tecnico['ticket_medio']:.2f}".replace('.', ','),
            tecnico['ultima_ordem'].strftime('%d/%m/%Y') if tecnico['ultima_ordem'] else '-',
            f"{taxa_conclusao:.1f}%".replace('.', ',')
        ])
    
    # Adicionar resumo
    writer.writerow([])
    writer.writerow(['Resumo', '', '', '', '', '', ''])
    writer.writerow(['Total de Técnicos', total_tecnicos, '', '', '', '', ''])
    writer.writerow(['Total de Ordens', total_ordens, '', '', '', '', ''])
    writer.writerow(['Valor Total (R$)', f"{total_valor:.2f}".replace('.', ','), '', '', '', '', ''])
    
    return response
